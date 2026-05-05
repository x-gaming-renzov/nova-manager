"""Verify simulation engine against the actual Excel file.

Reads each Excel sheet, extracts assumptions via the seed script's
extract_assumptions(), runs compute_simulation(), and compares every
output metric against the Excel's pre-computed values.

This is the definitive source-of-truth test — if these pass, the engine
matches the spreadsheet.

Requires: openpyxl, the Excel file at docs/Battlin_GTM_KPI_Simulator_v3.xlsx
"""

import sys
from pathlib import Path

import pytest

# Add project root to path for script imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from nova_manager.components.simulations.engine import compute_simulation

# Reuse the seed script's Excel parsing utilities
from scripts.seed_gtm_kpis import (
    EXCEL_PATH,
    SHEET_ALIASES,
    MONTH_COLS,
    PERIOD_LABELS,
    extract_assumptions,
    extract_business_data,
    extract_expected_values,
)

try:
    import openpyxl
    EXCEL_AVAILABLE = openpyxl is not None and EXCEL_PATH.exists()
except ImportError:
    EXCEL_AVAILABLE = False

skip_if_no_excel = pytest.mark.skipif(
    not EXCEL_AVAILABLE,
    reason="Excel file or openpyxl not available",
)

MONTHS = ["2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]


def _get_engine_values(rows, metric_name, dimension=""):
    """Extract monthly values from engine output."""
    filtered = [r for r in rows if r["metric_name"] == metric_name and r["dimension"] == dimension]
    by_period = {r["period_start"][:7]: r["value"] for r in filtered}
    return {m: by_period.get(m) for m in MONTHS if by_period.get(m) is not None}


def _get_excel_values(biz_data, metric_name, dimension=""):
    """Extract monthly values from Excel-extracted business data."""
    filtered = [r for r in biz_data if r["metric_name"] == metric_name and r["dimension"] == dimension]
    return {r["period_start"][:7]: r["value"] for r in filtered}


def _compare(engine_vals, excel_vals, label, tol=0.01):
    """Compare engine output against Excel values."""
    errors = []
    for month in sorted(set(engine_vals) | set(excel_vals)):
        eng = engine_vals.get(month)
        exc = excel_vals.get(month)
        if eng is None or exc is None:
            continue
        if exc == 0:
            if abs(eng) > 0.5:
                errors.append(f"{label} [{month}]: engine={eng}, excel=0")
        else:
            rel_err = abs(eng - exc) / abs(exc)
            if rel_err > tol:
                errors.append(f"{label} [{month}]: engine={eng:.4f}, excel={exc:.4f} (err={rel_err:.2%})")
    return errors


# Metrics to compare (metric_name, dimension)
METRICS_TO_VERIFY = [
    ("active_tos", ""),
    ("total_tournaments", ""),
    ("mau", ""),
    ("dau", ""),
    ("inorganic_players", ""),
    ("fill_rate", ""),
    ("total_player_slots", ""),
    ("avg_participants", ""),
    ("to_incentive_l1", "milestones"),
    ("to_incentive_l2", "leaderboard"),
    ("to_incentive_l3", "grand_prize"),
    ("to_incentive_total", ""),
    ("sponsored_credits", ""),
    ("supply_ua", ""),
    ("demand_ua", ""),
    ("ad_revenue", ""),
    ("sponsorship_revenue", ""),
    ("webshop_revenue", ""),
    ("total_revenue", ""),
    ("total_marketing_spend", ""),
]


@skip_if_no_excel
class TestExcelV2:
    """Verify engine against the James(v2) Excel sheet."""

    @pytest.fixture(autouse=True)
    def setup(self):
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["KPI Simulator_James(v2)"]
        self.assumptions = extract_assumptions(ws)
        self.engine_rows = compute_simulation(self.assumptions)
        self.excel_data = extract_business_data(ws)
        self.expected_kpis = extract_expected_values(ws)

    def test_all_metrics_match_excel(self):
        """Every derived metric matches the Excel's pre-computed values."""
        all_errors = []
        for metric_name, dimension in METRICS_TO_VERIFY:
            engine_vals = _get_engine_values(self.engine_rows, metric_name, dimension)
            excel_vals = _get_excel_values(self.excel_data, metric_name, dimension)
            errors = _compare(engine_vals, excel_vals, f"{metric_name}[{dimension}]" if dimension else metric_name)
            all_errors.extend(errors)

        if all_errors:
            pytest.fail(f"{len(all_errors)} mismatches:\n" + "\n".join(all_errors))

    def test_active_tos_cascade(self):
        engine_tos = _get_engine_values(self.engine_rows, "active_tos")
        excel_tos = _get_excel_values(self.excel_data, "active_tos")
        errors = _compare(engine_tos, excel_tos, "active_tos")
        assert not errors, "\n".join(errors)

    def test_total_tournaments(self):
        errors = _compare(
            _get_engine_values(self.engine_rows, "total_tournaments"),
            _get_excel_values(self.excel_data, "total_tournaments"),
            "total_tournaments",
        )
        assert not errors, "\n".join(errors)

    def test_supply_ua(self):
        errors = _compare(
            _get_engine_values(self.engine_rows, "supply_ua"),
            _get_excel_values(self.excel_data, "supply_ua"),
            "supply_ua",
        )
        assert not errors, "\n".join(errors)

    def test_total_revenue(self):
        errors = _compare(
            _get_engine_values(self.engine_rows, "total_revenue"),
            _get_excel_values(self.excel_data, "total_revenue"),
            "total_revenue",
        )
        assert not errors, "\n".join(errors)

    def test_total_marketing_spend(self):
        errors = _compare(
            _get_engine_values(self.engine_rows, "total_marketing_spend"),
            _get_excel_values(self.excel_data, "total_marketing_spend"),
            "total_marketing_spend",
        )
        assert not errors, "\n".join(errors)

    def test_kpi_cac_matches(self):
        """CAC = total_marketing_spend / mau — verify against Excel efficiency section."""
        spend = _get_engine_values(self.engine_rows, "total_marketing_spend")
        mau = _get_engine_values(self.engine_rows, "mau")
        expected = self.expected_kpis.get("cac", {})
        errors = []
        for month in sorted(spend):
            if month in mau and mau[month] > 0 and month in expected:
                computed_cac = spend[month] / mau[month]
                excel_cac = expected[month]
                rel_err = abs(computed_cac - excel_cac) / abs(excel_cac)
                if rel_err > 0.01:
                    errors.append(f"CAC [{month}]: {computed_cac:.4f} vs {excel_cac:.4f}")
        assert not errors, "\n".join(errors)

    def test_kpi_roas_matches(self):
        """ROAS = total_revenue / total_marketing_spend."""
        rev = _get_engine_values(self.engine_rows, "total_revenue")
        spend = _get_engine_values(self.engine_rows, "total_marketing_spend")
        expected = self.expected_kpis.get("roas", {})
        errors = []
        for month in sorted(rev):
            if month in spend and spend[month] > 0 and month in expected:
                computed = rev[month] / spend[month]
                excel = expected[month]
                rel_err = abs(computed - excel) / abs(excel)
                if rel_err > 0.01:
                    errors.append(f"ROAS [{month}]: {computed:.6f} vs {excel:.6f}")
        assert not errors, "\n".join(errors)


@skip_if_no_excel
class TestExcelNeutral:
    """Verify engine against the James(Neutral) sheet."""

    @pytest.fixture(autouse=True)
    def setup(self):
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["KPI Simulator_James(Neutral)"]
        self.assumptions = extract_assumptions(ws)
        self.engine_rows = compute_simulation(self.assumptions)
        self.excel_data = extract_business_data(ws)

    def test_all_metrics_match_excel(self):
        all_errors = []
        for metric_name, dimension in METRICS_TO_VERIFY:
            engine_vals = _get_engine_values(self.engine_rows, metric_name, dimension)
            excel_vals = _get_excel_values(self.excel_data, metric_name, dimension)
            errors = _compare(engine_vals, excel_vals, f"{metric_name}[{dimension}]" if dimension else metric_name)
            all_errors.extend(errors)
        if all_errors:
            pytest.fail(f"{len(all_errors)} mismatches:\n" + "\n".join(all_errors))

    def test_neutral_has_fewer_tos_than_v2(self):
        """Neutral scenario has fewer new TOs → lower active_tos."""
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        v2_assumptions = extract_assumptions(wb["KPI Simulator_James(v2)"])
        v2_rows = compute_simulation(v2_assumptions)
        v2_tos = _get_engine_values(v2_rows, "active_tos")
        neutral_tos = _get_engine_values(self.engine_rows, "active_tos")
        # After first month, v2 adds 200/month vs neutral 50/month
        for month in ["2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]:
            assert neutral_tos[month] < v2_tos[month], f"Neutral TOs should be < v2 in {month}"


@skip_if_no_excel
class TestExcelPositive:
    """Verify engine against the James(Positive) sheet."""

    @pytest.fixture(autouse=True)
    def setup(self):
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["KPI Simulator_James (Pos)"]
        self.assumptions = extract_assumptions(ws)
        self.engine_rows = compute_simulation(self.assumptions)
        self.excel_data = extract_business_data(ws)

    def test_all_metrics_match_excel(self):
        all_errors = []
        for metric_name, dimension in METRICS_TO_VERIFY:
            engine_vals = _get_engine_values(self.engine_rows, metric_name, dimension)
            excel_vals = _get_excel_values(self.excel_data, metric_name, dimension)
            errors = _compare(engine_vals, excel_vals, f"{metric_name}[{dimension}]" if dimension else metric_name)
            all_errors.extend(errors)
        if all_errors:
            pytest.fail(f"{len(all_errors)} mismatches:\n" + "\n".join(all_errors))


@skip_if_no_excel
class TestExcelAkshay:
    """Verify engine against the Akshay sheet."""

    @pytest.fixture(autouse=True)
    def setup(self):
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["KPI Simulator_Akshay"]
        self.assumptions = extract_assumptions(ws)
        self.engine_rows = compute_simulation(self.assumptions)
        self.excel_data = extract_business_data(ws)

    def test_all_metrics_match_excel(self):
        all_errors = []
        for metric_name, dimension in METRICS_TO_VERIFY:
            engine_vals = _get_engine_values(self.engine_rows, metric_name, dimension)
            excel_vals = _get_excel_values(self.excel_data, metric_name, dimension)
            errors = _compare(engine_vals, excel_vals, f"{metric_name}[{dimension}]" if dimension else metric_name)
            all_errors.extend(errors)
        if all_errors:
            pytest.fail(f"{len(all_errors)} mismatches:\n" + "\n".join(all_errors))

    def test_akshay_matches_v2(self):
        """Akshay sheet has identical data to v2 — engine outputs should match."""
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        v2_rows = compute_simulation(extract_assumptions(wb["KPI Simulator_James(v2)"]))
        for metric_name, dimension in METRICS_TO_VERIFY:
            v2_vals = _get_engine_values(v2_rows, metric_name, dimension)
            ak_vals = _get_engine_values(self.engine_rows, metric_name, dimension)
            for month in sorted(set(v2_vals) & set(ak_vals)):
                if v2_vals[month] == 0:
                    assert abs(ak_vals[month]) < 0.01
                else:
                    rel = abs(v2_vals[month] - ak_vals[month]) / abs(v2_vals[month])
                    assert rel < 0.01, f"{metric_name} [{month}]: v2={v2_vals[month]}, akshay={ak_vals[month]}"
