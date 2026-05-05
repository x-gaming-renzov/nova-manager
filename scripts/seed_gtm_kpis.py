#!/usr/bin/env python3
"""Seed GTM KPI metrics from the Battlin GTM KPI Simulator Excel.

Extracts business data from the Excel, ingests it via the API,
creates formula metric definitions, then validates each formula
by computing it and comparing against the Excel's expected values.

Usage:
  python -m scripts.seed_gtm_kpis --base-url http://localhost:8000
  python -m scripts.seed_gtm_kpis --dry-run          # just print payloads
  python -m scripts.seed_gtm_kpis --sheet "KPI Simulator_James(v2)"
"""

import argparse
import json
import math
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

EXCEL_PATH = Path(__file__).resolve().parent.parent / "docs" / "Battlin_GTM_KPI_Simulator_v3.xlsx"

# Month columns in the Excel (F=Jul, G=Aug, ..., K=Dec)
MONTH_COLS = {"F": "2026-07-01", "G": "2026-08-01", "H": "2026-09-01",
              "I": "2026-10-01", "J": "2026-11-01", "K": "2026-12-01"}

SHEET_ALIASES = {
    "v2": "KPI Simulator_James(v2)",
    "neutral": "KPI Simulator_James(Neutral)",
    "positive": "KPI Simulator_James (Pos)",
    "akshay": "KPI Simulator_Akshay",
}

# ── Excel row map per sheet ──────────────────────────────────────
# Each sheet has slightly different row numbers; these are for v2/Akshay.
# Neutral and Positive have offsets we detect dynamically.

def _find_row(ws, col_c_text: str, start: int = 1, end: int = 140,
              unit: str | None = None) -> int | None:
    """Find the row where column C contains the given text.

    Args:
        unit: If provided, also require column D to match this unit (e.g. '#', '$', '%').
    """
    for row in range(start, end + 1):
        val = ws[f"C{row}"].value
        if val and col_c_text.lower() in str(val).lower():
            if unit is not None:
                d_val = ws[f"D{row}"].value
                if d_val is None or str(d_val).strip() != unit:
                    continue
            return row
    return None


def _read_monthly(ws, row: int) -> dict[str, float]:
    """Read monthly values from columns F-K for a given row."""
    values = {}
    for col, period in MONTH_COLS.items():
        cell = ws[f"{col}{row}"]
        if cell.value is not None:
            try:
                values[period] = float(cell.value)
            except (ValueError, TypeError):
                pass
    return values


# ── Extract business data from one sheet ─────────────────────────

# Metrics to extract: (metric_name, search_text_in_col_C, dimension, unit_filter)
# We search by text to handle row number differences between sheets.
# unit_filter disambiguates when multiple rows match (e.g. "DAU" vs "DAU / MAU %").
METRICS_TO_EXTRACT = [
    # Supply
    ("active_tos",              "Active Tournament Organizers",   "",           "#"),
    ("total_tournaments",       "Total Tournaments / Month",      "",           None),
    ("to_incentive_l1",         "Layer 1 Total (USD)",            "milestones", None),
    ("to_incentive_l2",         "Layer 2 Total (USD)",            "leaderboard",None),
    ("to_incentive_l3",         "Layer 3 Total (USD)",            "grand_prize",None),
    ("to_incentive_total",      "TO Incentives Total (L1+L2+L3)", "",           None),
    ("sponsored_credits",       "Sponsored Credits (USD)",        "",           "$"),
    ("supply_ua",               "TOTAL SUPPLY UA (USD)",          "",           None),

    # Demand
    ("mau",                     "MAU",                            "",           "#"),
    ("dau",                     "DAU",                            "",           "#"),
    ("inorganic_players",       "Inorganic Players",              "",           "#"),
    ("demand_ua",               "Total Demand UA",                "",           "$"),

    # Engagement
    ("fill_rate",               "Avg Fill Rate",                  "",           "%"),
    ("total_player_slots",      "Total Player Slots",             "",           "#"),
    ("avg_participants",        "Avg Participants",               "",           "#"),

    # Revenue
    ("ad_revenue",              "Monthly Ad Revenue (USD)",       "",                None),
    ("sponsorship_revenue",     "Sponsorship Revenue (USD)",      "",                None),
    ("webshop_revenue",         "Webshop Revenue (USD)",          "",                None),
    ("total_revenue",           "TOTAL REVENUE (USD)",            "",                None),

    # Marketing budget by channel
    ("marketing_spend",         "TO Incentives (L1+L2+L3) → derived", "to_empowerment", None),
    ("marketing_spend",         "Agency Retainer",                "social_channel",  None),
    ("marketing_spend",         "Sponsored Credits → derived",    "prize_pools",     None),
    ("marketing_spend",         "Campus Ambassadors",             "college_activation", None),
    ("marketing_spend",         "BGMI Streamers",                 "kol_influencer",  None),
    ("marketing_spend",         "Search & Store Optimization",    "seo_aso",         None),
    ("marketing_spend",         "UC Codes & Rewards",             "bgmi_giveaways",  None),
    ("marketing_spend",         "Press Releases",                 "pr_media",        None),
    ("marketing_spend",         "BATTLIN BGMI Open",              "marquee_tourney", None),
    ("total_marketing_spend",   "TOTAL MARKETING SPEND",          "",                None),
]

# Expected values for validation (search text, metric_name_for_display)
EXPECTED_METRICS = [
    ("CAC (Total Mkt / MAU)",       "cac"),
    ("Supply CAC",                  "supply_cac"),
    ("ROAS (Revenue / Spend)",      "roas"),
    ("ARPU (Revenue / MAU)",        "arpu"),
    ("Cost per Tournament",         "cost_per_tournament"),
    ("Revenue per Tournament",      "revenue_per_tournament"),
    ("Net Margin (Rev - Mkt Spend)","net_margin"),
    ("Margin %",                    "margin_pct"),
]


def extract_business_data(ws) -> list[dict]:
    """Extract all business data rows from one Excel sheet."""
    rows = []

    for metric_name, search_text, dimension, unit_filter in METRICS_TO_EXTRACT:
        row_num = _find_row(ws, search_text, unit=unit_filter)
        if row_num is None:
            print(f"  WARN  Could not find row for '{search_text}'")
            continue

        monthly = _read_monthly(ws, row_num)
        for period, value in monthly.items():
            rows.append({
                "metric_name": metric_name,
                "dimension": dimension,
                "value": value,
                "period_start": f"{period}T00:00:00Z",
                "currency": "USD",
            })

    return rows


def extract_expected_values(ws) -> dict[str, dict[str, float]]:
    """Extract expected metric values from the Efficiency section for validation."""
    expected = {}
    for search_text, name in EXPECTED_METRICS:
        row_num = _find_row(ws, search_text)
        if row_num is None:
            continue
        monthly = _read_monthly(ws, row_num)
        expected[name] = monthly
    return expected


# ── Formula metric definitions ───────────────────────────────────

FORMULA_METRICS = [
    {
        "name": "CAC",
        "description": "Customer Acquisition Cost = Total Marketing Spend / MAU",
        "type": "formula",
        "config": {
            "time_range": {"start": "2026-07-01 00:00:00", "end": "2027-01-01 00:00:00"},
            "granularity": "monthly",
            "group_by": [],
            "operands": {
                "spend": {
                    "type": "operational",
                    "config": {"metric_name": "total_marketing_spend", "aggregation": "sum"},
                },
                "mau": {
                    "type": "operational",
                    "config": {"metric_name": "mau", "aggregation": "sum"},
                },
            },
            "expression": "spend / mau",
        },
    },
    {
        "name": "ROAS",
        "description": "Return on Ad Spend = Total Revenue / Total Marketing Spend",
        "type": "formula",
        "config": {
            "time_range": {"start": "2026-07-01 00:00:00", "end": "2027-01-01 00:00:00"},
            "granularity": "monthly",
            "group_by": [],
            "operands": {
                "revenue": {
                    "type": "operational",
                    "config": {"metric_name": "total_revenue", "aggregation": "sum"},
                },
                "spend": {
                    "type": "operational",
                    "config": {"metric_name": "total_marketing_spend", "aggregation": "sum"},
                },
            },
            "expression": "revenue / spend",
        },
    },
    {
        "name": "ARPU",
        "description": "Average Revenue Per User = Total Revenue / MAU",
        "type": "formula",
        "config": {
            "time_range": {"start": "2026-07-01 00:00:00", "end": "2027-01-01 00:00:00"},
            "granularity": "monthly",
            "group_by": [],
            "operands": {
                "revenue": {
                    "type": "operational",
                    "config": {"metric_name": "total_revenue", "aggregation": "sum"},
                },
                "mau": {
                    "type": "operational",
                    "config": {"metric_name": "mau", "aggregation": "sum"},
                },
            },
            "expression": "revenue / mau",
        },
    },
    {
        "name": "Net Margin",
        "description": "Net Margin = Total Revenue - Total Marketing Spend",
        "type": "formula",
        "config": {
            "time_range": {"start": "2026-07-01 00:00:00", "end": "2027-01-01 00:00:00"},
            "granularity": "monthly",
            "group_by": [],
            "operands": {
                "revenue": {
                    "type": "operational",
                    "config": {"metric_name": "total_revenue", "aggregation": "sum"},
                },
                "spend": {
                    "type": "operational",
                    "config": {"metric_name": "total_marketing_spend", "aggregation": "sum"},
                },
            },
            "expression": "revenue - spend",
        },
    },
    {
        "name": "Margin %",
        "description": "Margin Percentage = (Revenue - Spend) / Revenue",
        "type": "formula",
        "config": {
            "time_range": {"start": "2026-07-01 00:00:00", "end": "2027-01-01 00:00:00"},
            "granularity": "monthly",
            "group_by": [],
            "operands": {
                "revenue": {
                    "type": "operational",
                    "config": {"metric_name": "total_revenue", "aggregation": "sum"},
                },
                "spend": {
                    "type": "operational",
                    "config": {"metric_name": "total_marketing_spend", "aggregation": "sum"},
                },
            },
            "expression": "(revenue - spend) / revenue",
        },
    },
    {
        "name": "Supply CAC",
        "description": "Supply-side CAC = Total Supply UA / Active TOs",
        "type": "formula",
        "config": {
            "time_range": {"start": "2026-07-01 00:00:00", "end": "2027-01-01 00:00:00"},
            "granularity": "monthly",
            "group_by": [],
            "operands": {
                "supply_ua": {
                    "type": "operational",
                    "config": {"metric_name": "supply_ua", "aggregation": "sum"},
                },
                "active_tos": {
                    "type": "operational",
                    "config": {"metric_name": "active_tos", "aggregation": "sum"},
                },
            },
            "expression": "supply_ua / active_tos",
        },
    },
    {
        "name": "Demand CAC",
        "description": "Demand-side CAC = Total Demand UA / Inorganic Players",
        "type": "formula",
        "config": {
            "time_range": {"start": "2026-07-01 00:00:00", "end": "2027-01-01 00:00:00"},
            "granularity": "monthly",
            "group_by": [],
            "operands": {
                "demand_ua": {
                    "type": "operational",
                    "config": {"metric_name": "demand_ua", "aggregation": "sum"},
                },
                "inorganic_players": {
                    "type": "operational",
                    "config": {"metric_name": "inorganic_players", "aggregation": "sum"},
                },
            },
            "expression": "demand_ua / inorganic_players",
        },
    },
    {
        "name": "Cost per Tournament",
        "description": "Cost per Tournament = Total Marketing Spend / Total Tournaments",
        "type": "formula",
        "config": {
            "time_range": {"start": "2026-07-01 00:00:00", "end": "2027-01-01 00:00:00"},
            "granularity": "monthly",
            "group_by": [],
            "operands": {
                "spend": {
                    "type": "operational",
                    "config": {"metric_name": "total_marketing_spend", "aggregation": "sum"},
                },
                "tournaments": {
                    "type": "operational",
                    "config": {"metric_name": "total_tournaments", "aggregation": "sum"},
                },
            },
            "expression": "spend / tournaments",
        },
    },
    {
        "name": "Revenue per Tournament",
        "description": "Revenue per Tournament = Total Revenue / Total Tournaments",
        "type": "formula",
        "config": {
            "time_range": {"start": "2026-07-01 00:00:00", "end": "2027-01-01 00:00:00"},
            "granularity": "monthly",
            "group_by": [],
            "operands": {
                "revenue": {
                    "type": "operational",
                    "config": {"metric_name": "total_revenue", "aggregation": "sum"},
                },
                "tournaments": {
                    "type": "operational",
                    "config": {"metric_name": "total_tournaments", "aggregation": "sum"},
                },
            },
            "expression": "revenue / tournaments",
        },
    },
]

# Map formula names to expected-value keys
FORMULA_TO_EXPECTED = {
    "CAC": "cac",
    "ROAS": "roas",
    "ARPU": "arpu",
    "Net Margin": "net_margin",
    "Margin %": "margin_pct",
    "Supply CAC": "supply_cac",
    "Cost per Tournament": "cost_per_tournament",
    "Revenue per Tournament": "revenue_per_tournament",
}

PERIOD_LABELS = {
    "2026-07-01": "Jul", "2026-08-01": "Aug", "2026-09-01": "Sep",
    "2026-10-01": "Oct", "2026-11-01": "Nov", "2026-12-01": "Dec",
}


# ── Main ─────────────────────────────────────────────────────────

def _compute_config_with_scenario(fm: dict, scenario_id: str) -> dict:
    """Return a compute config with scenario_id injected into operational operands."""
    config = json.loads(json.dumps(fm["config"]))  # deep copy
    for operand in config.get("operands", {}).values():
        if operand["type"] == "operational":
            operand["config"]["scenario_id"] = scenario_id
    return config


def _ingest_scenario(api: str, headers: dict, biz_data: list[dict], scenario_id: str) -> int:
    """Ingest business data rows under a scenario_id. Returns count ingested."""
    batch_size = 50
    total = 0
    for i in range(0, len(biz_data), batch_size):
        batch = biz_data[i:i + batch_size]
        r = requests.post(f"{api}/metrics/business-data/",
                          headers=headers,
                          json={"data": batch, "scenario_id": scenario_id})
        if r.status_code == 200:
            total += r.json().get("count", 0)
        else:
            print(f"  FAIL  Ingest batch {i}: {r.status_code} — {r.text[:200]}")
            return -1
    return total


def _validate_scenario(api: str, headers: dict, scenario_id: str,
                       expected: dict, formulas: list[dict]) -> tuple[int, int]:
    """Compute each formula for a scenario and compare against expected values.
    Returns (pass_count, fail_count)."""
    pass_count = 0
    fail_count = 0

    for fm in formulas:
        name = fm["name"]
        expected_key = FORMULA_TO_EXPECTED.get(name)
        expected_monthly = expected.get(expected_key, {})

        compute_config = _compute_config_with_scenario(fm, scenario_id)
        r = requests.post(f"{api}/metrics/compute/", headers=headers, json={
            "type": fm["type"],
            "config": compute_config,
        })

        print(f"\n  {name}  ({fm['config']['expression']})")
        if r.status_code != 200:
            print(f"    FAIL  Compute returned {r.status_code}: {r.text[:200]}")
            fail_count += 1
            continue

        data = r.json()
        if not data:
            print(f"    FAIL  No data returned")
            fail_count += 1
            continue

        metric_passed = True
        for row in sorted(data, key=lambda x: str(x.get("period", ""))):
            period_str = str(row.get("period", ""))[:10]
            computed = row.get("value")
            label = PERIOD_LABELS.get(period_str, period_str)
            exp = expected_monthly.get(period_str)

            if computed is None:
                print(f"    {label}:  computed=NULL, expected={_fmt(exp)}")
                continue

            if exp is not None:
                if exp == 0:
                    match = abs(computed) < 0.01
                else:
                    match = abs(computed - exp) / abs(exp) < 0.01
                status = "PASS" if match else "FAIL"
                if not match:
                    metric_passed = False
                print(f"    {label}:  computed={_fmt(computed)}, expected={_fmt(exp)}  [{status}]")
            else:
                print(f"    {label}:  computed={_fmt(computed)}")

        if metric_passed:
            pass_count += 1
        else:
            fail_count += 1

    return pass_count, fail_count


def main():
    parser = argparse.ArgumentParser(description="Seed GTM KPI metrics from Excel")
    parser.add_argument("--base-url", default="http://localhost:8000")
    parser.add_argument("--sheet", default="v2",
                        help="Sheet alias (v2, neutral, positive, akshay) or full name")
    parser.add_argument("--all-scenarios", action="store_true",
                        help="Ingest and validate all 4 Excel scenarios under one account")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print payloads without calling the API")
    args = parser.parse_args()

    base = args.base_url.rstrip("/")
    api = f"{base}/api/v1"

    # Determine which sheets to process
    if args.all_scenarios:
        sheets_to_run = list(SHEET_ALIASES.items())  # [(alias, full_name), ...]
    else:
        alias = args.sheet
        full_name = SHEET_ALIASES.get(alias, alias)
        sheets_to_run = [(alias, full_name)]

    # ── Load Excel ───────────────────────────────────────────
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Extract data for all requested sheets
    scenario_data = {}  # {alias: {biz_data, expected, sheet_name}}
    for alias, sheet_name in sheets_to_run:
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            sys.exit(1)
        ws = wb[sheet_name]
        biz_data = extract_business_data(ws)
        expected = extract_expected_values(ws)
        scenario_data[alias] = {"biz_data": biz_data, "expected": expected, "sheet_name": sheet_name}

    # ── Phase 1: Show extracted data ─────────────────────────
    for alias, sdata in scenario_data.items():
        scenario_id = f"scenario_{alias}"
        biz_data = sdata["biz_data"]

        print(f"\n{'=' * 60}")
        print(f"  Phase 1: Extract — {sdata['sheet_name']}  (scenario: {scenario_id})")
        print(f"{'=' * 60}")
        print(f"\nExtracted {len(biz_data)} business data rows")

        by_metric = {}
        for row in biz_data:
            key = row["metric_name"]
            if row["dimension"]:
                key += f" [{row['dimension']}]"
            by_metric.setdefault(key, []).append(row)

        for metric_key, rows in sorted(by_metric.items()):
            vals = ", ".join(
                f"{PERIOD_LABELS.get(r['period_start'][:10], r['period_start'][:7])}={_fmt(r['value'])}"
                for r in sorted(rows, key=lambda x: x["period_start"])
            )
            print(f"  {metric_key}: {vals}")

    if args.dry_run:
        sample = list(scenario_data.values())[0]["biz_data"]
        print(f"\n[DRY RUN] Sample payload (scenario_{sheets_to_run[0][0]}):")
        print(json.dumps({"scenario_id": f"scenario_{sheets_to_run[0][0]}", "data": sample[:3]}, indent=2, default=str))
        print(f"  ... ({len(sample)} rows)")
        print("\n[DRY RUN] Phase 2 & 3 skipped.")
        return

    # ── Setup: Fresh Account ─────────────────────────────────
    import uuid
    run_id = uuid.uuid4().hex[:8]

    print(f"\n{'=' * 60}")
    print(f"  Setup: Fresh Account (run: {run_id})")
    print(f"{'=' * 60}")

    auth = _fresh_account_setup(api, run_id)
    if not auth:
        sys.exit(1)
    headers = auth["headers"]

    # ── Ingest each scenario ─────────────────────────────────
    for alias, sdata in scenario_data.items():
        scenario_id = f"scenario_{alias}"
        biz_data = sdata["biz_data"]

        print(f"\n  Ingesting {len(biz_data)} rows as '{scenario_id}'...")
        count = _ingest_scenario(api, headers, biz_data, scenario_id)
        if count < 0:
            sys.exit(1)
        print(f"  OK    Ingested {count} rows")

    # ── Schema discovery ─────────────────────────────────────
    r = requests.get(f"{api}/metrics/business-data/schema/", headers=headers)
    if r.status_code == 200:
        schema = r.json()
        scenarios_found = sorted({row.get("scenario_id", "?") for row in schema})
        metric_count = len({row["metric_name"] for row in schema})
        print(f"\n  Schema: {metric_count} metric names across {len(scenarios_found)} scenarios: {scenarios_found}")

    # ── Phase 2: Create formula metrics (scenario-agnostic) ──
    print(f"\n{'=' * 60}")
    print("  Phase 2: Create Formula Metric Definitions (no scenario)")
    print(f"{'=' * 60}")

    saved_metrics = {}
    for fm in FORMULA_METRICS:
        print(f"\n  {fm['name']}: {fm['config']['expression']}")
        r = requests.post(f"{api}/metrics/", headers=headers, json=fm)
        if r.status_code == 200:
            metric = r.json()
            saved_metrics[fm["name"]] = metric["pid"]
            print(f"    SAVED  pid={metric['pid']}")
        else:
            print(f"    FAIL   {r.status_code}: {r.text[:200]}")

    # ── Phase 3: Validate each scenario ──────────────────────
    total_pass = 0
    total_fail = 0

    for alias, sdata in scenario_data.items():
        scenario_id = f"scenario_{alias}"
        expected = sdata["expected"]

        print(f"\n{'=' * 60}")
        print(f"  Phase 3: Validate — {scenario_id}")
        print(f"{'=' * 60}")

        p, f = _validate_scenario(api, headers, scenario_id, expected, FORMULA_METRICS)
        total_pass += p
        total_fail += f

    # ── Phase 3b: Verify saved metric retrieval ──────────────
    if saved_metrics:
        print(f"\n{'─' * 60}")
        print(f"  Phase 3b: Verify saved metric retrieval")
        print(f"{'─' * 60}")
        for name, pid in list(saved_metrics.items())[:3]:
            r = requests.get(f"{api}/metrics/{pid}/", headers=headers)
            if r.status_code == 200:
                m = r.json()
                has_scenario = any(
                    "scenario_id" in op.get("config", {})
                    for op in m["config"].get("operands", {}).values()
                )
                print(f"  OK    {name} (pid={pid}): type={m['type']}, scenario_in_def={has_scenario}")
            else:
                print(f"  FAIL  {name} (pid={pid}): {r.status_code}")

    # ── Summary ──────────────────────────────────────────────
    n_scenarios = len(scenario_data)
    print(f"\n{'=' * 60}")
    print(f"  Results: {total_pass}/{total_pass + total_fail} formulas match Excel")
    print(f"  Scenarios tested: {n_scenarios} ({', '.join(f'scenario_{a}' for a in scenario_data)})")
    print(f"  Saved metrics: {len(saved_metrics)} (scenario-agnostic)")
    print(f"{'=' * 60}")

    if saved_metrics:
        print(f"\n  Metric IDs:")
        for name, pid in saved_metrics.items():
            print(f"    {name}: {pid}")

    sys.exit(0 if total_fail == 0 else 1)


def _fmt(v) -> str:
    """Format a value for display."""
    if v is None:
        return "N/A"
    if isinstance(v, str):
        return v
    if abs(v) >= 1000:
        return f"{v:,.2f}"
    if abs(v) >= 1:
        return f"{v:.4f}"
    if abs(v) >= 0.001:
        return f"{v:.6f}"
    return f"{v:.8f}"


def _fresh_account_setup(api: str, run_id: str) -> dict | None:
    """Register a fresh account, create an app, and return auth headers.

    Returns headers dict with Authorization + Content-Type, ready for API calls.
    """
    email = f"gtm_seed_{run_id}@test.com"
    password = "SeedTest123!"
    company = f"Battlin Seed {run_id}"

    # 1. Register
    print(f"\n  Registering: {email}")
    r = requests.post(f"{api}/auth/register", json={
        "email": email,
        "password": password,
        "name": f"GTM Seed {run_id}",
        "company": company,
    })
    if r.status_code != 200:
        print(f"  FAIL  Register: {r.status_code} — {r.text[:200]}")
        return None

    data = r.json()
    token = data["access_token"]
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    print(f"  OK    Registered")

    # 2. Create app
    print(f"  Creating app: Battlin GTM")
    r = requests.post(f"{api}/auth/apps", headers=headers, json={
        "name": f"Battlin GTM {run_id}",
        "description": "GTM KPI Simulator seed test",
    })
    if r.status_code != 200:
        print(f"  FAIL  Create app: {r.status_code} — {r.text[:200]}")
        return None

    app_data = r.json()
    # App creation returns new tokens with app context baked in
    token = app_data.get("access_token", token)
    headers["Authorization"] = f"Bearer {token}"
    app_id = app_data.get("app", {}).get("id", "?")
    print(f"  OK    App created: {app_id}")

    # 3. Get SDK credentials (needed for any event tracking)
    r = requests.get(f"{api}/auth/sdk-credentials", headers=headers)
    sdk_key = None
    if r.status_code == 200:
        sdk_key = r.json().get("sdk_api_key")
        print(f"  OK    SDK key obtained")

    return {"headers": headers, "sdk_key": sdk_key, "app_id": app_id}


if __name__ == "__main__":
    main()
